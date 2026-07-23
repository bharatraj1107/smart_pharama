import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()

# Sheet 1: Employee Master
master = wb.active
master.title = 'Employee Master'
headers_master = ['Emp No', 'Employee Name', 'Department', 'Shift Timing', 'Contact No']
master.append(headers_master)
for col in range(1, len(headers_master) + 1):
    master.column_dimensions[get_column_letter(col)].width = 20

# Sheet 2: Attendance
att = wb.create_sheet('Attendance')
headers_att = ['Emp No', 'Employee Name', 'Date', 'Day', 'Status', 'Time In', 'Time Out', 'OD/Extra Hours', 'Working Hours', 'Overtime', 'Day-wise Total Hours', 'Remarks']
att.append(headers_att)
for col in range(1, len(headers_att) + 1):
    att.column_dimensions[get_column_letter(col)].width = 18

# Sheet 3: Advance Money
adv = wb.create_sheet('Advance Money')
headers_adv = ['Emp No', 'Employee Name', 'Advance Amount', 'Reason/Purpose', 'Date Given', 'Time Given', 'Given By', 'Repayment Mode', 'Status', 'Amount Repaid', 'Balance Due', 'Remarks']
adv.append(headers_adv)
for col in range(1, len(headers_adv) + 1):
    adv.column_dimensions[get_column_letter(col)].width = 18

# Sheet 4: Dashboard Summary
dash = wb.create_sheet('Dashboard Summary')
headers_dash = ['Emp No', 'Employee Name', 'Total Working Hours', 'Total Overtime', 'Total Advance Outstanding']
dash.append(headers_dash)
for col in range(1, len(headers_dash) + 1):
    dash.column_dimensions[get_column_letter(col)].width = 22

# Freeze panes
master.freeze_panes = 'A2'
att.freeze_panes = 'A2'
adv.freeze_panes = 'A2'
dash.freeze_panes = 'A2'

# Data validation lists
status_vals = ['Present', 'Absent', 'Half Day', 'OD', 'Leave', 'WFH']
reason_vals = ['Salary Advance', 'Emergency', 'Travel', 'Other']
repay_vals = ['Salary Deduction', 'Cash', 'Bank Transfer']

status_dv = DataValidation(type='list', formula1='"' + ','.join(status_vals) + '"', allow_blank=True)
reason_dv = DataValidation(type='list', formula1='"' + ','.join(reason_vals) + '"', allow_blank=True)
repay_dv = DataValidation(type='list', formula1='"' + ','.join(repay_vals) + '"', allow_blank=True)
att.add_data_validation(status_dv)
adv.add_data_validation(reason_dv)
adv.add_data_validation(repay_dv)

for row in range(2, 202):
    status_dv.add(f'E{row}')
    reason_dv.add(f'D{row}')
    repay_dv.add(f'H{row}')

# Header style helper

def apply_header_style(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill('solid', fgColor='FFCCCCCC')
        cell.alignment = Alignment(horizontal='center', vertical='center')

apply_header_style(master)
apply_header_style(att)
apply_header_style(adv)
apply_header_style(dash)

# Attendance formulas
for row in range(2, 202):
    att[f'B{row}'] = f'=IF(A{row}="", "", IFERROR(VLOOKUP(A{row}, "Employee Master"!$A:$B, 2, FALSE), ""))'
    att[f'D{row}'] = f'=IF(C{row}="", "", TEXT(C{row}, "ddd"))'
    att[f'I{row}'] = f'=IF(OR(E{row}="Absent", E{row}="Leave", E{row}=""), "", IF(AND(F{row}>0, G{row}>0), G{row}-F{row}, ""))'
    att[f'J{row}'] = f'=IF(I{row}="", "", MAX(0, I{row}-TIME(9,0,0)))'
    att[f'K{row}'] = f'=IF(I{row}="", "", I{row} + IF(H{row}="", 0, H{row}))'

# Advance Money formulas
for row in range(2, 202):
    adv[f'B{row}'] = f'=IF(A{row}="", "", IFERROR(VLOOKUP(A{row}, "Employee Master"!$A:$B, 2, FALSE), ""))'
    adv[f'E{row}'] = f'=IF(A{row}="", "", TODAY())'
    adv[f'F{row}'] = f'=IF(A{row}="", "", NOW())'
    adv[f'K{row}'] = f'=IF(C{row}="", "", C{row}-J{row})'
    adv[f'I{row}'] = f'=IF(K{row}="", "", IF(K{row}=0, "Repaid", IF(K{row}<C{row}, "Partially Repaid", "Repaid")))'

# Dashboard formulas
for row in range(2, 102):
    dash[f'B{row}'] = f'=IF(A{row}="", "", IFERROR(VLOOKUP(A{row}, "Employee Master"!$A:$B, 2, FALSE), ""))'
    dash[f'C{row}'] = f'=IF(A{row}="", "", SUMIFS(Attendance!$I:$I, Attendance!$A:$A, A{row}, Attendance!$I:$I, ">0"))'
    dash[f'D{row}'] = f'=IF(A{row}="", "", SUMIFS(Attendance!$J:$J, Attendance!$A:$A, A{row}, Attendance!$J:$J, ">0"))'
    dash[f'E{row}'] = f'=IF(A{row}="", "", SUMIFS(\'Advance Money\'!$K:$K, \'Advance Money\'!$A:$A, A{row}))'

# Summary block labels in Attendance and Advance Money
att['M1'] = 'Attendance Summary'
att['M2'] = 'Total Present'
att['M3'] = 'Total Absent'
att['M4'] = 'Total OD Hours'
att['M5'] = 'Total Working Hours'
att['M6'] = 'Total Overtime'

adv['M1'] = 'Advance Summary'
adv['M2'] = 'Total Advance Taken'
adv['M3'] = 'Total Repaid'
adv['M4'] = 'Total Outstanding Balance'

# Status conditional formatting for Attendance
att.conditional_formatting.add('E2:E201',
    FormulaRule(formula=['$E2="Present"'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
att.conditional_formatting.add('E2:E201',
    FormulaRule(formula=['$E2="Absent"'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))
att.conditional_formatting.add('E2:E201',
    FormulaRule(formula=['OR($E2="Half Day", $E2="OD")'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))

# Status conditional formatting for Advance Money
adv.conditional_formatting.add('I2:I201',
    FormulaRule(formula=['$I2="Repaid"'], fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')))
adv.conditional_formatting.add('I2:I201',
    FormulaRule(formula=['$I2="Partially Repaid"'], fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')))
adv.conditional_formatting.add('I2:I201',
    FormulaRule(formula=['$I2="Pending"'], fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')))

# Save workbook
wb.save('attendance_advance_workbook.xlsx')
print('Workbook created: attendance_advance_workbook.xlsx')
